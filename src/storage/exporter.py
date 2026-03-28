"""
src/storage/exporter.py

Exports products to CSV, Excel, and JSON.

Author : M-D (Mohamed Darwish)
"""

import csv
import json
from datetime import datetime
from pathlib import Path

from src.models.product import Product
from src.utils.logger import get_logger


logger = get_logger(__name__)


class DataExporter:

    def __init__(self, output_dir: str = "data/processed"):
        self.out = Path(output_dir)
        self.out.mkdir(parents=True, exist_ok=True)

    def to_csv(self, products: list[Product], filename: str) -> Path:
        if not products:
            raise ValueError("No products to export.")
        path  = self.out / filename
        rows  = [p.to_dict() for p in products]
        fields = list(rows[0].keys())
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
        logger.info(f"CSV: {path} ({len(rows)} rows)")
        return path

    def to_excel(self, products: list[Product], filename: str) -> Path:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        if not products:
            raise ValueError("No products to export.")
        path  = self.out / filename
        rows  = [p.to_dict() for p in products]
        fields = list(rows[0].keys())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        hfill = PatternFill("solid", fgColor="1A1A2E")
        hfont = Font(color="E0E0E0", bold=True, name="Calibri", size=10)

        for ci, f in enumerate(fields, 1):
            cell = ws.cell(row=1, column=ci, value=f)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            for ci, f in enumerate(fields, 1):
                ws.cell(row=ri, column=ci, value=row.get(f))

        for ci, f in enumerate(fields, 1):
            max_len = max(len(f), *[len(str(r.get(f, "") or "")) for r in rows[:50]])
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 48)

        ws.freeze_panes = "A2"
        wb.save(path)
        logger.info(f"Excel: {path} ({len(rows)} rows)")
        return path

    def to_json(self, products: list[Product], filename: str) -> Path:
        path = self.out / filename
        data = {
            "meta": {
                "exported_at":    datetime.utcnow().isoformat(),
                "total_products": len(products),
                "tool":           "M-D Web Scraper",
            },
            "products": [p.to_dict() for p in products],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        logger.info(f"JSON: {path} ({len(products)} records)")
        return path

    def export_all(self, products: list[Product], stem: str) -> dict[str, Path]:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        base = f"{stem}_{ts}"
        return {
            "csv":   self.to_csv(products,   f"{base}.csv"),
            "excel": self.to_excel(products, f"{base}.xlsx"),
            "json":  self.to_json(products,  f"{base}.json"),
        }
